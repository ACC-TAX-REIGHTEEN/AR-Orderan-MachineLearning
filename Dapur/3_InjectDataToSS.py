import configparser
from collections import Counter, defaultdict
from datetime import datetime, timedelta
import os
import re
import time
import warnings
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from rapidfuzz import fuzz, process
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)


def load_config():
    config = configparser.ConfigParser()
    config.read("config.conf")
    return config


def get_general_config(config):
    noo_keywords = []
    if config.has_section("GENERAL"):
        gen_new_cust = config.get("GENERAL", "gen_new_cust", fallback="").strip()
        if gen_new_cust:
            noo_keywords = [
                k.strip().lower() for k in gen_new_cust.split(",") if k.strip()
            ]
    return noo_keywords


def get_custom_rules(config):
    group_keywords = []
    if config.has_section("GROUP_KEYWORDS"):
        for _, val in config.items("GROUP_KEYWORDS"):
            if val.strip():
                group_keywords.append(val.strip().lower())

    branch_rules = {}
    if config.has_section("BRANCH_RULES"):
        for key, val in config.items("BRANCH_RULES"):
            if key.strip() and val.strip():
                branch_rules[key.strip().lower()] = val.strip()

    return group_keywords, branch_rules


def get_product_config(config, suffix):
    if not config.has_section("AR"):
        return "", "", "", "", "", ""

    url = config.get("AR", f"ar_url_{suffix}", fallback="").strip()
    sheet = config.get("AR", f"ar_sheet_{suffix}", fallback="Sheet1").strip()
    key_col = config.get(
        "AR", f"ar_key_col_{suffix}", fallback="Nama Customer dan Kota"
    ).strip()
    prod_key_col = config.get(
        "AR", f"ar_prod_key_col_{suffix}", fallback=""
    ).strip()
    target_col = config.get(
        "AR", f"ar_target_col_{suffix}", fallback="Nominal Nota Belum Lunas"
    ).strip()
    key_filter = config.get(
        "AR", f"ar_key_filter_{suffix}", fallback=suffix.upper()
    ).strip()

    return url, sheet, key_col, prod_key_col, target_col, key_filter


def bersihkan_teks(teks):
    if pd.isna(teks):
        return ""
    t = str(teks).lower()
    t = re.sub(r"[^\w\s]", " ", t)
    return " ".join(t.split())


def extract_dynamic_generic_words(
    df_ar, col_pelanggan, col_kontak, df_ml=None, df_fb=None, threshold_ratio=0.015
):
    token_counter = Counter()
    unique_texts = set()

    if col_pelanggan and col_pelanggan in df_ar.columns:
        unique_texts.update(df_ar[col_pelanggan].dropna().astype(str).unique())
    if col_kontak and col_kontak in df_ar.columns:
        unique_texts.update(df_ar[col_kontak].dropna().astype(str).unique())

    if df_ml is not None:
        for col in ["Nama Customer dan Kota", "Hasil_Nama_Rekomendasi"]:
            if col in df_ml.columns:
                unique_texts.update(df_ml[col].dropna().astype(str).unique())

    if df_fb is not None:
        for col in ["KETERANGAN", "NAMA"]:
            if col in df_fb.columns:
                unique_texts.update(df_fb[col].dropna().astype(str).unique())

    total_records = len(unique_texts)
    if total_records == 0:
        return set()

    for text in unique_texts:
        tokens = set(bersihkan_teks(text).split())
        for token in tokens:
            if len(token) > 1:
                token_counter[token] += 1

    dynamic_detected = {
        token
        for token, count in token_counter.items()
        if (count / total_records) >= threshold_ratio
    }

    return dynamic_detected


def validate_ml_match(target_clean, candidate_key, dynamic_stopwords=None):
    if dynamic_stopwords is None:
        dynamic_stopwords = set()

    target_tokens = bersihkan_teks(target_clean).split()
    candidate_tokens = set(bersihkan_teks(candidate_key).split())

    core_tokens = [
        t for t in target_tokens
        if t not in dynamic_stopwords and len(t) > 1
    ]

    if not core_tokens:
        core_tokens = [t for t in target_tokens if len(t) > 1]

    for c_token in core_tokens:
        has_match = False
        if c_token in candidate_tokens:
            has_match = True
        else:
            for cand_t in candidate_tokens:
                if len(c_token) >= 3 and len(cand_t) >= 3:
                    if cand_t.startswith(c_token) or c_token.startswith(cand_t):
                        has_match = True
                        break
        if not has_match:
            return False

    return True


def format_idr(value):
    if (
        pd.isna(value)
        or str(value).strip().lower() == "nan"
        or str(value).strip() == ""
    ):
        return "0"
    try:
        val_float = float(value)
        return f"{val_float:,.0f}".replace(",", ".")
    except (ValueError, TypeError):
        return str(value)


def parse_order_date(date_str):
    try:
        dt = pd.to_datetime(date_str, errors="coerce", format="mixed")
        if pd.notna(dt):
            return dt.strftime("%d/%m/%Y")
    except Exception:
        pass
    return str(date_str)


def format_excel_date(date_val):
    if pd.isna(date_val):
        return ""
    if isinstance(date_val, datetime) or hasattr(date_val, "strftime"):
        return date_val.strftime("%d %b %Y")
    return str(date_val)


def parse_max_jt_date(jt_str):
    if pd.isna(jt_str) or not str(jt_str).strip():
        return None

    clean_str = re.sub(r"(?i)jt\s*bg", "", str(jt_str)).strip()
    if not clean_str:
        return None

    parsed_dates = []
    groups = clean_str.split("&")

    for group in groups:
        group = group.strip()
        if not group:
            continue

        parts = group.split("/")
        if len(parts) == 3:
            days_str, month_str, year_str = (
                parts[0].strip(),
                parts[1].strip(),
                parts[2].strip(),
            )

            if len(year_str) == 2:
                year_str = f"20{year_str}"

            days = [d.strip() for d in days_str.split(",") if d.strip()]
            for d in days:
                try:
                    dt = datetime(
                        int(year_str), int(month_str), int(d)
                    ).date()
                    parsed_dates.append(dt)
                except ValueError:
                    continue

    if parsed_dates:
        return max(parsed_dates)
    return None


def standardize_code(
    code, depo_prefixes="SL|YY|MKS|MGL|PW|PWT|PLU|SG|SMG|TGL|PA|KDI"
):
    if pd.isna(code):
        return ""
    if isinstance(code, float) and code.is_integer():
        code = int(code)
    s = str(code).strip().upper()
    s = re.sub(r"\s*-\s*", "-", s)
    pattern = rf"^({depo_prefixes})\s*(\d+)"
    s = re.sub(pattern, r"\1-\2", s)
    s = re.sub(r"(\d+)([A-Z])$", r"\1 \2", s)
    return s


def load_minifs_mapping(minifs_file="Minifs_temp.xlsx"):
    code_to_all_codes = defaultdict(set)
    if not os.path.exists(minifs_file):
        return code_to_all_codes

    try:
        df_m = pd.read_excel(minifs_file)
        df_m.columns = df_m.columns.str.strip()

        col_min = [
            c
            for c in df_m.columns
            if "min" in c.lower() and "pelanggan" in c.lower()
        ]
        col_nopel = [
            c
            for c in df_m.columns
            if c.lower() == "no. pelanggan" or c.lower() == "kode pelanggan"
        ]

        if col_min and col_nopel:
            c_min = col_min[0]
            c_nop = col_nopel[0]

            def clean_c(v):
                if pd.isna(v):
                    return ""
                if isinstance(v, (int, float)):
                    return str(int(v))
                s = str(v).strip()
                if s.endswith(".0"):
                    s = s[:-2]
                return s.upper()

            df_m["Min_Clean"] = df_m[c_min].apply(clean_c)
            df_m["Nopel_Clean"] = df_m[c_nop].apply(clean_c)

            min_to_group = defaultdict(set)
            for _, r in df_m.iterrows():
                m_val = r["Min_Clean"]
                n_val = r["Nopel_Clean"]
                if m_val:
                    min_to_group[m_val].add(m_val)
                if n_val:
                    if m_val:
                        min_to_group[m_val].add(n_val)
                    else:
                        min_to_group[n_val].add(n_val)

            for m_val, group_set in min_to_group.items():
                for code_item in group_set:
                    code_to_all_codes[code_item].update(group_set)

    except Exception as e:
        print(f"--> [WARNING MINIFS]: Gagal memuat Minifs_temp.xlsx: {e}")

    return code_to_all_codes


def read_excel_auto_header(file_path, sheet_name=0, target_column=""):
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    target_clean = str(target_column).strip().upper()

    for idx, row in df_raw.iterrows():
        row_cleaned = [str(val).strip().upper() for val in row.dropna()]
        if target_clean in row_cleaned:
            df_clean = df_raw.iloc[idx + 1 :].copy()
            df_clean.columns = df_raw.iloc[idx].astype(str).str.strip()
            return df_clean.reset_index(drop=True)

    raise KeyError(
        f"Kolom target '{target_column}' tidak ditemukan pada file '{file_path}'"
    )


def simpan_ar_ke_local_excel(user_ar_rows, nama_pelanggan_resmi, config):
    if not user_ar_rows:
        return

    save_dir = config.get(
        "DIR", "save_to_local", fallback=r"E:\ADM IRC AND ZN\2026\AGU\\"
    ).strip()

    if not os.path.exists(save_dir):
        try:
            os.makedirs(save_dir, exist_ok=True)
        except Exception as e_dir:
            print(f"--> [ERROR] Gagal membuat direktori {save_dir}: {e_dir}")
            return

    def get_val(row_dict, key_candidates, default=""):
        for k in key_candidates:
            if k in row_dict and pd.notna(row_dict[k]):
                return str(row_dict[k]).strip()
        return default

    first_row = user_ar_rows[0]
    nama_kontak = get_val(first_row, ["Nama Kontak", "nama kontak"])
    nama_penjual = get_val(first_row, ["Nama Penjual", "nama penjual"])

    indo_months_short = {
        1: "JAN",
        2: "PEB",
        3: "MAR",
        4: "APR",
        5: "MEI",
        6: "JUN",
        7: "JUL",
        8: "AGU",
        9: "SEP",
        10: "OKT",
        11: "NOP",
        12: "DES",
    }
    now = datetime.now()
    tgl_str = (
        f"{now.day:02d} {indo_months_short[now.month]} {str(now.year)[-2:]}"
    )

    raw_file_name = (
        f"{nama_pelanggan_resmi}, {nama_kontak}, {nama_penjual} , {tgl_str}"
    )
    clean_file_name = re.sub(r'[\\/:*?"<>|]', "-", raw_file_name)
    clean_file_name = re.sub(r"\s+", " ", clean_file_name).strip()

    full_save_path = os.path.join(save_dir, f"{clean_file_name}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    font_header = Font(name="Arial", size=10, bold=True, color="000266")
    font_data = Font(name="Arial", size=10)
    font_total = Font(name="Arial", size=10, bold=True, color="000266")

    headers_layout = {
        1: "No. Faktur",
        2: "Tgl Faktur",
        4: "Jatuh Tempo",
        6: "Nilai Faktur",
        7: "Sisa Piutang",
        8: "Umur JT",
        9: "Nama Pelanggan",
        10: "Nama Penjual",
        11: "Nama Kontak",
    }

    for col_idx, col_name in headers_layout.items():
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = font_header

    row_start = 2
    for idx, r in enumerate(user_ar_rows, start=row_start):
        no_faktur = get_val(r, ["No. Faktur", "no. faktur", "no faktur"])
        tgl_faktur = format_excel_date(r.get("Tgl Faktur"))
        jt_faktur = format_excel_date(r.get("Jatuh Tempo"))

        try:
            nilai_faktur = (
                float(r.get("Nilai Faktur", 0))
                if pd.notna(r.get("Nilai Faktur"))
                else 0.0
            )
        except Exception:
            nilai_faktur = 0.0

        try:
            sisa_piutang = (
                float(r.get("Sisa Piutang", 0))
                if pd.notna(r.get("Sisa Piutang"))
                else 0.0
            )
        except Exception:
            sisa_piutang = 0.0

        umur_jt = get_val(r, ["Umur JT", "umur jt"])
        c_pelanggan = get_val(
            r,
            ["Nama Pelanggan", "nama pelanggan"],
            default=nama_pelanggan_resmi,
        )
        c_penjual = get_val(r, ["Nama Penjual", "nama penjual"])
        c_kontak = get_val(r, ["Nama Kontak", "nama kontak"])

        ws.cell(row=idx, column=1, value=no_faktur).font = font_data
        ws.cell(row=idx, column=2, value=tgl_faktur).font = font_data
        ws.cell(row=idx, column=4, value=jt_faktur).font = font_data

        cell_nf = ws.cell(row=idx, column=6, value=nilai_faktur)
        cell_nf.font = font_data
        cell_nf.number_format = "#,##0"

        cell_sp = ws.cell(row=idx, column=7, value=sisa_piutang)
        cell_sp.font = font_data
        cell_sp.number_format = "#,##0"

        ws.cell(row=idx, column=8, value=umur_jt).font = font_data
        ws.cell(row=idx, column=9, value=c_pelanggan).font = font_data
        ws.cell(row=idx, column=10, value=c_penjual).font = font_data
        ws.cell(row=idx, column=11, value=c_kontak).font = font_data

    last_data_row = row_start + len(user_ar_rows) - 1
    total_row = last_data_row + 1

    cell_tot_f = ws.cell(
        row=total_row, column=6, value=f"=SUM(F2:F{last_data_row})"
    )
    cell_tot_f.font = font_total
    cell_tot_f.number_format = "#,##0"

    cell_tot_g = ws.cell(
        row=total_row, column=7, value=f"=SUM(G2:G{last_data_row})"
    )
    cell_tot_g.font = font_total
    cell_tot_g.number_format = "#,##0"

    try:
        wb.save(full_save_path)
        print(f"--> [LOKAL BERHASIL] File disave: {clean_file_name}.xlsx")
    except Exception as e_save:
        print(f"--> [LOKAL GAGAL] Tidak dapat menyimpan file: {e_save}")


def preload_all_data_to_memory(flag_fraud, ar_key_filter, config):
    print(
        f"--> RAM PRELOAD Memuat data ke RAM dengan Filter Produk: '{ar_key_filter}'..."
    )

    depo_config = config.get(
        "MAP",
        "depo",
        fallback="SL|YY|MKS|MGL|PW|PWT|PLU|SG|SMG|TGL|PA|KDI",
    ).strip()
    minifs_mapping = load_minifs_mapping("Minifs_temp.xlsx")

    df_ml_temp = None
    ml_dict = {}
    ml_list = []
    if os.path.exists("Hasil_Latihan_temp.xlsx"):
        df_ml_temp = pd.read_excel("Hasil_Latihan_temp.xlsx")
        df_ml_temp.columns = df_ml_temp.columns.str.strip()
        if (
            "Nama Customer dan Kota" in df_ml_temp.columns
            and "Hasil_Nama_Rekomendasi" in df_ml_temp.columns
        ):
            for _, r in df_ml_temp.iterrows():
                k = bersihkan_teks(r["Nama Customer dan Kota"])
                v = str(r["Hasil_Nama_Rekomendasi"]).strip()
                if (
                    k
                    and v
                    and v.upper()
                    not in ["TIDAK DITEMUKAN", "FAILED", "NAN", "NONE", ""]
                ):
                    ml_dict[k] = v
                    if k not in ml_list:
                        ml_list.append(k)

    df_fb_temp = None
    fb_dict = {}
    fb_list = []
    if os.path.exists("FBackCust_temp.xlsx"):
        df_fb_temp = pd.read_excel("FBackCust_temp.xlsx")
        df_fb_temp.columns = df_fb_temp.columns.str.strip()

        if "NO." in df_fb_temp.columns:
            df_fb_temp["NO."] = pd.to_numeric(df_fb_temp["NO."], errors="coerce")
            df_fb_temp = df_fb_temp.sort_values(by="NO.", ascending=True)
            df_fb_temp = df_fb_temp.drop_duplicates(subset=["KETERANGAN"], keep="last")

        if "KETERANGAN" in df_fb_temp.columns and "NAMA" in df_fb_temp.columns:
            for _, row in df_fb_temp.iterrows():
                ket = bersihkan_teks(row["KETERANGAN"])
                nama = str(row["NAMA"]).strip()
                if ket and nama:
                    fb_dict[ket] = nama
                    fb_list.append(ket)

    ar_memory = defaultdict(list)
    dynamic_stopwords = set()

    if os.path.exists("ARClean_temp.xlsx"):
        df_ar = read_excel_auto_header(
            "ARClean_temp.xlsx", sheet_name=0, target_column="Nama Pelanggan"
        )

        col_map = {str(c).strip().lower(): c for c in df_ar.columns}
        col_penjual = col_map.get("nama penjual")
        col_kontak = col_map.get("nama kontak")
        col_pelanggan = col_map.get("nama pelanggan")
        col_raw_code = col_map.get("no. pelanggan") or col_map.get(
            "kode pelanggan"
        )

        dynamic_stopwords = extract_dynamic_generic_words(
            df_ar=df_ar,
            col_pelanggan=col_pelanggan,
            col_kontak=col_kontak,
            df_ml=df_ml_temp,
            df_fb=df_fb_temp,
            threshold_ratio=0.015,
        )

        if col_penjual:
            df_ar = df_ar[
                df_ar[col_penjual]
                .astype(str)
                .str.contains("SR", case=False, na=False)
            ]

        if ar_key_filter and str(ar_key_filter).strip():
            filter_str = str(ar_key_filter).strip()

            cond_kontak = (
                df_ar[col_kontak]
                .astype(str)
                .str.contains(filter_str, case=False, na=False)
                if col_kontak
                else pd.Series(False, index=df_ar.index)
            )

            cond_penjual = (
                df_ar[col_penjual]
                .astype(str)
                .str.contains(filter_str, case=False, na=False)
                if col_penjual
                else pd.Series(False, index=df_ar.index)
            )

            cond_pelanggan = (
                df_ar[col_pelanggan]
                .astype(str)
                .str.contains(filter_str, case=False, na=False)
                if col_pelanggan
                else pd.Series(False, index=df_ar.index)
            )

            df_ar = df_ar[cond_kontak | cond_penjual | cond_pelanggan]

        if flag_fraud == "No" and col_penjual:
            df_ar = df_ar[
                ~df_ar[col_penjual]
                .astype(str)
                .str.contains("FRAUD", case=False, na=False)
            ]

        col_faktur = col_map.get("no. faktur") or col_map.get("no faktur")
        if col_faktur:
            df_ar = df_ar[
                df_ar[col_faktur].notna()
                & (df_ar[col_faktur].astype(str).str.strip() != "")
                & (
                    df_ar[col_faktur].astype(str).str.strip().str.lower()
                    != "nan"
                )
            ]

        indo_months = {
            "mei": "may",
            "ags": "aug",
            "agt": "aug",
            "agu": "aug",
            "okt": "oct",
            "nop": "nov",
            "des": "dec",
            "peb": "feb",
        }

        def parse_date_sort(val):
            if pd.isna(val):
                return pd.NaT
            val_str = str(val).lower().strip()
            for indo, eng in indo_months.items():
                if indo in val_str:
                    val_str = val_str.replace(indo, eng)
                    break
            return pd.to_datetime(val_str, errors="coerce", format="mixed")

        col_tgl_faktur = col_map.get("tgl faktur")
        if col_tgl_faktur:
            df_ar["Temp_Sort_Date"] = df_ar[col_tgl_faktur].apply(
                parse_date_sort
            )

        sort_cols = [
            c
            for c in ["Temp_Sort_Date", col_faktur]
            if c and c in df_ar.columns
        ]
        if sort_cols:
            df_ar = df_ar.sort_values(by=sort_cols, ascending=[True, True])

        for _, r in df_ar.iterrows():
            r_dict = r.to_dict()
            p_key = (
                bersihkan_teks(r.get(col_pelanggan, ""))
                if col_pelanggan
                else ""
            )
            k_key = bersihkan_teks(r.get(col_kontak, "")) if col_kontak else ""

            if p_key:
                ar_memory[p_key].append(r_dict)
            if k_key and k_key != p_key:
                ar_memory[k_key].append(r_dict)

            if col_raw_code and pd.notna(r.get(col_raw_code)):
                raw_c = str(r.get(col_raw_code)).strip()
                if raw_c and raw_c.lower() != "nan":
                    clean_raw = raw_c[:-2] if raw_c.endswith(".0") else raw_c
                    std_c = standardize_code(clean_raw, depo_config)

                    ar_memory[f"CODE_{clean_raw.lower()}"].append(r_dict)
                    if std_c.lower() != clean_raw.lower():
                        ar_memory[f"CODE_{std_c.lower()}"].append(r_dict)

    print(
        f"--> [DYNAMIC STOPWORDS] Terdeteksi {len(dynamic_stopwords)} kata generik/wilayah secara otomatis dari korpus data."
    )
    print("--> RAM PRELOAD SELESAI Data terfilter presisi!\n")
    return (
        ml_dict,
        ml_list,
        fb_dict,
        fb_list,
        ar_memory,
        minifs_mapping,
        depo_config,
        dynamic_stopwords,
    )


def resolve_target_name_fast(
    raw_key,
    ml_dict,
    ml_list,
    fb_dict,
    fb_list,
    cache_resolver,
    branch_rules,
    dynamic_stopwords=None,
):
    if dynamic_stopwords is None:
        dynamic_stopwords = set()

    raw_clean = bersihkan_teks(raw_key)
    if not raw_clean:
        return str(raw_key).strip()

    if raw_clean in cache_resolver:
        return cache_resolver[raw_clean]

    for rule_key, target_name in branch_rules.items():
        if "|" in rule_key:
            parent_kw, branch_kw = rule_key.split("|", 1)
            parent_kw = parent_kw.strip()
            branch_kw = branch_kw.strip()

            if parent_kw in raw_clean and branch_kw in raw_clean:
                cache_resolver[raw_clean] = target_name
                return target_name
        else:
            if rule_key in raw_clean:
                cache_resolver[raw_clean] = target_name
                return target_name

    if raw_clean in fb_dict:
        res = fb_dict[raw_clean]
        cache_resolver[raw_clean] = res
        return res

    if raw_clean in ml_dict:
        res = ml_dict[raw_clean]
        cache_resolver[raw_clean] = res
        return res

    if ml_list:
        matches_ml = process.extract(
            query=raw_clean,
            choices=ml_list,
            scorer=fuzz.token_set_ratio,
            score_cutoff=65.0,
            limit=10,
        )
        if not matches_ml:
            matches_ml = process.extract(
                query=raw_clean,
                choices=ml_list,
                scorer=fuzz.WRatio,
                score_cutoff=65.0,
                limit=10,
            )
        if matches_ml:
            for match_item in matches_ml:
                candidate_key = match_item[0]
                if validate_ml_match(raw_clean, candidate_key, dynamic_stopwords):
                    res = ml_dict[candidate_key]
                    cache_resolver[raw_clean] = res
                    return res

    if fb_list:
        matches_fb = process.extract(
            query=raw_clean,
            choices=list(fb_dict.keys()),
            scorer=fuzz.token_set_ratio,
            score_cutoff=65.0,
            limit=10,
        )
        if not matches_fb:
            matches_fb = process.extract(
                query=raw_clean,
                choices=list(fb_dict.keys()),
                scorer=fuzz.WRatio,
                score_cutoff=65.0,
                limit=10,
            )
        if matches_fb:
            for match_item in matches_fb:
                candidate_key = match_item[0]
                if validate_ml_match(raw_clean, candidate_key, dynamic_stopwords):
                    res = fb_dict[candidate_key]
                    cache_resolver[raw_clean] = res
                    return res

    res = str(raw_key).strip()
    cache_resolver[raw_clean] = res
    return res


def get_ar_rows_fast(
    target_clean,
    raw_key_clean,
    ar_memory,
    cache_ar_lookup,
    group_keywords,
    minifs_mapping,
    depo_config,
    ar_key_filter="",
    dynamic_stopwords=None,
):
    if dynamic_stopwords is None:
        dynamic_stopwords = set()

    cache_key = f"{raw_key_clean}|{target_clean}"
    if cache_key in cache_ar_lookup:
        return cache_ar_lookup[cache_key]

    matched_group = None
    for g_kw in group_keywords:
        if g_kw in raw_key_clean or g_kw in target_clean:
            matched_group = g_kw
            break

    if matched_group:
        g_cache_key = f"GROUP_{matched_group}_{ar_key_filter}_{cache_key}"
        if g_cache_key in cache_ar_lookup:
            return cache_ar_lookup[g_cache_key]

        combined_rows = []
        seen_invoices = set()
        prod_filter_clean = str(ar_key_filter).lower().strip()

        for ar_key, rows in ar_memory.items():
            if matched_group in ar_key:
                for r in rows:
                    kontak_val = str(r.get("Nama Kontak", "")).lower()
                    penjual_val = str(r.get("Nama Penjual", "")).lower()

                    if matched_group in kontak_val:
                        if not prod_filter_clean or (
                            prod_filter_clean in kontak_val
                            or prod_filter_clean in penjual_val
                        ):
                            inv_no = str(r.get("No. Faktur", ""))
                            if inv_no not in seen_invoices:
                                seen_invoices.add(inv_no)
                                combined_rows.append(r)

        cache_ar_lookup[g_cache_key] = combined_rows
        return combined_rows

    raw_codes = [
        k.strip()
        for k in raw_key_clean.replace("nopel:", "").split("&")
        if k.strip()
    ]
    target_codes_set = set()
    for r_code in raw_codes:
        code_upper = r_code.upper()
        std_c = standardize_code(r_code, depo_config)
        target_codes_set.add(code_upper.lower())
        target_codes_set.add(std_c.lower())

        if code_upper in minifs_mapping:
            for m_code in minifs_mapping[code_upper]:
                target_codes_set.add(m_code.lower())
        if std_c in minifs_mapping:
            for m_code in minifs_mapping[std_c]:
                target_codes_set.add(m_code.lower())

    combined_code_rows = []
    seen_code_invoices = set()
    for c_key in target_codes_set:
        if f"CODE_{c_key}" in ar_memory:
            for r in ar_memory[f"CODE_{c_key}"]:
                inv_no = str(r.get("No. Faktur", ""))
                if inv_no not in seen_code_invoices:
                    seen_code_invoices.add(inv_no)
                    combined_code_rows.append(r)

    if combined_code_rows:
        cache_ar_lookup[cache_key] = combined_code_rows
        return combined_code_rows

    ar_keys = [k for k in ar_memory.keys() if not k.startswith("CODE_")]
    is_ml_mapped = bool(target_clean and target_clean != raw_key_clean)

    if is_ml_mapped:
        matched_rows = []
        seen_invoices = set()

        if target_clean in ar_memory and validate_ml_match(
            target_clean, target_clean, dynamic_stopwords
        ):
            for r in ar_memory[target_clean]:
                inv_no = str(r.get("No. Faktur", ""))
                if inv_no not in seen_invoices:
                    seen_invoices.add(inv_no)
                    matched_rows.append(r)
            if matched_rows:
                cache_ar_lookup[cache_key] = matched_rows
                return matched_rows

        target_tokens = [t for t in target_clean.split() if len(t) > 0]
        if target_tokens:
            matching_keys = [
                k
                for k in ar_keys
                if all(
                    re.search(rf"\b{re.escape(token)}\b", k)
                    for token in target_tokens
                )
                and validate_ml_match(target_clean, k, dynamic_stopwords)
            ]
            if matching_keys:
                for k in matching_keys:
                    for r in ar_memory[k]:
                        inv_no = str(r.get("No. Faktur", ""))
                        if inv_no not in seen_invoices:
                            seen_invoices.add(inv_no)
                            matched_rows.append(r)
                if matched_rows:
                    cache_ar_lookup[cache_key] = matched_rows
                    return matched_rows

        if ar_keys:
            matches = process.extract(
                query=target_clean,
                choices=ar_keys,
                scorer=fuzz.WRatio,
                score_cutoff=85.0,
                limit=10,
            )
            if not matches:
                matches = process.extract(
                    query=target_clean,
                    choices=ar_keys,
                    scorer=fuzz.token_set_ratio,
                    score_cutoff=85.0,
                    limit=10,
                )
            if matches:
                for m in matches:
                    if validate_ml_match(target_clean, m[0], dynamic_stopwords):
                        for r in ar_memory[m[0]]:
                            inv_no = str(r.get("No. Faktur", ""))
                            if inv_no not in seen_invoices:
                                seen_invoices.add(inv_no)
                                matched_rows.append(r)
                if matched_rows:
                    cache_ar_lookup[cache_key] = matched_rows
                    return matched_rows

        cache_ar_lookup[cache_key] = []
        return []

    matched_rows = []
    seen_invoices = set()

    query_tokens = [t for t in raw_key_clean.split() if len(t) > 0]
    if query_tokens:
        matching_keys = [
            k
            for k in ar_keys
            if all(
                re.search(rf"\b{re.escape(token)}\b", k)
                for token in query_tokens
            )
            and validate_ml_match(raw_key_clean, k, dynamic_stopwords)
        ]
        if matching_keys:
            for k in matching_keys:
                for r in ar_memory[k]:
                    inv_no = str(r.get("No. Faktur", ""))
                    if inv_no not in seen_invoices:
                        seen_invoices.add(inv_no)
                        matched_rows.append(r)
            if matched_rows:
                cache_ar_lookup[cache_key] = matched_rows
                return matched_rows

    if ar_keys and raw_key_clean:
        matches = process.extract(
            query=raw_key_clean,
            choices=ar_keys,
            scorer=fuzz.WRatio,
            score_cutoff=85.0,
            limit=10,
        )
        if not matches:
            matches = process.extract(
                query=raw_key_clean,
                choices=ar_keys,
                scorer=fuzz.token_set_ratio,
                score_cutoff=85.0,
                limit=10,
            )
        if matches:
            for m in matches:
                if validate_ml_match(raw_key_clean, m[0], dynamic_stopwords):
                    for r in ar_memory[m[0]]:
                        inv_no = str(r.get("No. Faktur", ""))
                        if inv_no not in seen_invoices:
                            seen_invoices.add(inv_no)
                            matched_rows.append(r)
            if matched_rows:
                cache_ar_lookup[cache_key] = matched_rows
                return matched_rows

    cache_ar_lookup[cache_key] = []
    return []


def run_ar_process():
    print(
        f"--> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} Memulai sinkronisasi data AR..."
    )

    config = load_config()
    group_keywords, branch_rules = get_custom_rules(config)

    flag_fraud = config.get("AR", "ar_data_fraud", fallback="No").strip()
    flag_codecus = config.get("AR", "ar_data_codecus", fallback="Ya").strip()
    flag_namecus = config.get("AR", "ar_data_namecus", fallback="Ya").strip()
    fallback_prod = config.get("AR", "ar_data_prod", fallback="PCMO").strip()
    flag_dt_order = config.get("AR", "ar_data_dt_order", fallback="Ya").strip()
    flag_calc = config.get("AR", "ar_data_calc", fallback="Ya").strip()

    flag_inv_numb = config.get("AR", "ar_data_inv_numb", fallback="No").strip()
    flag_inv_dt = config.get("AR", "ar_data_inv_dt", fallback="Ya").strip()
    flag_inv_due = config.get("AR", "ar_data_inv_due", fallback="No").strip()
    flag_inv_val = config.get("AR", "ar_data_inv_val", fallback="Ya").strip()
    flag_inv_orig = config.get("AR", "ar_data_inv_orig", fallback="Ya").strip()
    flag_inv_ar = config.get("AR", "ar_data_inv_ar", fallback="No").strip()
    flag_inv_pay = config.get("AR", "ar_data_inv_pay", fallback="Ya").strip()
    flag_giro = config.get("AR", "ar_data_giro", fallback="Ya").strip()
    flag_age = config.get("AR", "ar_data_age", fallback="Ya").strip()

    noo_keywords = get_general_config(config)
    daftar_produk = ["irc", "zn"]

    for prod_suffix in daftar_produk:
        (
            ar_url,
            ar_sheet_name,
            ar_key_col_name,
            ar_prod_key_col_name,
            ar_target_col_name,
            ar_key_filter,
        ) = get_product_config(config, prod_suffix)

        if not ar_url:
            print(
                f"--> MELEWATI PRODUK {prod_suffix.upper()} URL (ar_url_{prod_suffix}) tidak diisi di config.conf."
            )
            continue

        print(
            "--> =================================================================="
        )
        print(
            f"--> MEMPROSES PRODUK: {prod_suffix.upper()} (Filter: {ar_key_filter})"
        )
        print(
            "--> =================================================================="
        )

        (
            ml_dict,
            ml_list,
            fb_dict,
            fb_list,
            ar_memory,
            minifs_mapping,
            depo_config,
            dynamic_stopwords,
        ) = preload_all_data_to_memory(flag_fraud, ar_key_filter, config)

        cache_resolver = {}
        cache_ar_lookup = {}

        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(
            "credentials.json", scopes=scope
        )
        client = gspread.authorize(creds)

        try:
            ss = client.open_by_url(ar_url)
            wks = ss.worksheet(ar_sheet_name)
            sheet_id = wks.id
        except Exception as e_conn:
            print(
                f"--> Gagal membuka Google Sheet {prod_suffix.upper()}: {e_conn}"
            )
            continue

        all_rows = wks.get_all_values()
        if not all_rows:
            print(f"--> Sheet '{ar_sheet_name}' kosong.")
            continue

        header = all_rows[0]
        header_clean = [str(h).strip().upper() for h in header]

        try:
            key_col_idx = header_clean.index(ar_key_col_name.strip().upper())
            target_col_idx = header_clean.index(
                ar_target_col_name.strip().upper()
            )

            prod_key_clean = ar_prod_key_col_name.strip().upper()
            prod_col_idx = (
                header_clean.index(prod_key_clean)
                if prod_key_clean in header_clean
                else None
            )
        except ValueError as e:
            print(
                f"--> Kolom target di Google Sheets tidak ditemukan: {e}"
            )
            continue

        requests = []
        current_date = datetime.now().date()
        total_diisi = 0

        for row_idx, row in enumerate(all_rows[1:], start=2):
            if key_col_idx >= len(row):
                continue

            if target_col_idx < len(row) and row[target_col_idx].strip() != "":
                continue

            raw_key = row[key_col_idx]
            if not str(raw_key).strip():
                continue

            raw_key_str = str(raw_key).strip().lower()
            raw_order_dt = row[0] if len(row) > 0 else ""
            formatted_order_dt = parse_order_date(raw_order_dt)

            product_val = fallback_prod
            if (
                prod_col_idx is not None
                and prod_col_idx < len(row)
                and row[prod_col_idx].strip()
            ):
                product_val = row[prod_col_idx].strip()

            is_noo = False
            for kw in noo_keywords:
                if raw_key_str.startswith(kw) or (kw in raw_key_str):
                    is_noo = True
                    break

            if is_noo:
                note_lines = [
                    f"{raw_key}\t{product_val}",
                    f"Tanggal Order: {formatted_order_dt}",
                    "",
                    "========================================",
                    "RINGKASAN PERFORMA PIUTANG",
                    "========================================",
                    "Piutang\t\t\t\t\t :  0",
                    "Total Faktur Aktif (Inv) :  0",
                    "",
                    "========================================",
                    "DAFTAR RINCIAN FAKTUR AKTIF",
                    "========================================",
                    "Outlet Baru (NOO)",
                ]
                final_note_text = "\n".join(note_lines)

                req_item = {
                    "updateCells": {
                        "rows": [
                            {
                                "values": [
                                    {
                                        "userEnteredValue": {
                                            "stringValue": "0"
                                            if flag_calc == "Ya"
                                            else ""
                                        },
                                        "note": final_note_text,
                                    }
                                ]
                            }
                        ],
                        "fields": "userEnteredValue,note",
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row_idx - 1,
                            "endRowIndex": row_idx,
                            "startColumnIndex": target_col_idx,
                            "endColumnIndex": target_col_idx + 1,
                        },
                    }
                }
                requests.append(req_item)
                total_diisi += 1
                continue

            raw_key_clean = bersihkan_teks(raw_key)

            nama_target_resmi = resolve_target_name_fast(
                raw_key,
                ml_dict,
                ml_list,
                fb_dict,
                fb_list,
                cache_resolver,
                branch_rules,
                dynamic_stopwords,
            )
            target_clean = bersihkan_teks(nama_target_resmi)

            user_ar_rows = get_ar_rows_fast(
                target_clean,
                raw_key_clean,
                ar_memory,
                cache_ar_lookup,
                group_keywords,
                minifs_mapping,
                depo_config,
                ar_key_filter,
                dynamic_stopwords,
            )

            if user_ar_rows:

                def ambil_tanggal_sort(r):
                    tgl = r.get("Temp_Sort_Date")
                    if pd.notna(tgl):
                        return tgl
                    return pd.Timestamp.min

                user_ar_rows = sorted(user_ar_rows, key=ambil_tanggal_sort)

                simpan_ar_ke_local_excel(
                    user_ar_rows, nama_target_resmi, config
                )

            total_sisa_piutang = sum(
                [
                    float(r.get("Sisa Piutang", 0))
                    for r in user_ar_rows
                    if pd.notna(r.get("Sisa Piutang"))
                ]
            )
            formatted_piutang_val = format_idr(total_sisa_piutang)

            note_lines = []
            header_line_parts = []

            if flag_codecus == "Ya":
                header_line_parts.append(str(raw_key))
            if flag_namecus == "Ya":
                header_line_parts.append(nama_target_resmi)

            header_line_parts.append(product_val)
            note_lines.append("\t".join([p for p in header_line_parts if p]))

            if flag_dt_order == "Ya":
                note_lines.append(f"Tanggal Order: {formatted_order_dt}")

            note_lines.append("")
            note_lines.append("========================================")
            note_lines.append("RINGKASAN PERFORMA PIUTANG")
            note_lines.append("========================================")

            if flag_calc == "Ya":
                note_lines.append(
                    f"Piutang\t\t\t\t\t :  {formatted_piutang_val} "
                )
            if flag_inv_val == "Ya":
                note_lines.append(
                    f"Total Faktur Aktif (Inv) :  {len(user_ar_rows)} "
                )

            note_lines.append("")
            note_lines.append("========================================")
            note_lines.append("DAFTAR RINCIAN FAKTUR AKTIF")
            note_lines.append("========================================")

            if not user_ar_rows:
                note_lines.append("Data tidak ditemukan")
            else:
                prev_month_year = None
                indo_months = {
                    "mei": "may",
                    "ags": "aug",
                    "agt": "aug",
                    "agu": "aug",
                    "okt": "oct",
                    "nop": "nov",
                    "des": "dec",
                    "peb": "feb",
                }

                for inv_row in user_ar_rows:
                    tgl_dt = inv_row.get("Temp_Sort_Date")
                    curr_month_year = None

                    if pd.notna(tgl_dt) and hasattr(tgl_dt, "month"):
                        curr_month_year = (tgl_dt.year, tgl_dt.month)
                    else:
                        tgl_raw = str(inv_row.get("Tgl Faktur", "")).lower()
                        for indo, eng in indo_months.items():
                            if indo in tgl_raw:
                                tgl_raw = tgl_raw.replace(indo, eng)
                                break
                        tgl_parsed = pd.to_datetime(
                            tgl_raw, errors="coerce", format="mixed"
                        )
                        if pd.notna(tgl_parsed):
                            curr_month_year = (
                                tgl_parsed.year,
                                tgl_parsed.month,
                            )

                    if (
                        prev_month_year is not None
                        and curr_month_year is not None
                        and curr_month_year != prev_month_year
                    ):
                        note_lines.append("")

                    if curr_month_year is not None:
                        prev_month_year = curr_month_year

                    inv_part = []

                    if flag_inv_numb == "Ya":
                        inv_part.append(str(inv_row.get("No. Faktur", "")))
                    if flag_inv_dt == "Ya":
                        inv_part.append(
                            format_excel_date(inv_row.get("Tgl Faktur"))
                        )
                    if flag_inv_due == "Ya":
                        inv_part.append(
                            format_excel_date(inv_row.get("Jatuh Tempo"))
                        )
                    if flag_inv_orig == "Ya":
                        inv_part.append(
                            format_idr(inv_row.get("Nilai Faktur", 0))
                        )
                    if flag_inv_ar == "Ya":
                        inv_part.append(
                            format_idr(inv_row.get("Sisa Piutang", 0))
                        )

                    if flag_inv_pay == "Ya":
                        try:
                            nilai_faktur = (
                                float(inv_row.get("Nilai Faktur", 0))
                                if pd.notna(inv_row.get("Nilai Faktur"))
                                else 0.0
                            )
                            sisa_piutang = (
                                float(inv_row.get("Sisa Piutang", 0))
                                if pd.notna(inv_row.get("Sisa Piutang"))
                                else 0.0
                            )
                            ttp_byr_val = nilai_faktur - sisa_piutang
                            if ttp_byr_val > 0:
                                inv_part.append(
                                    f"Ttp Byr: {format_idr(ttp_byr_val)}"
                                )
                        except (ValueError, TypeError):
                            pass

                    if flag_age == "Ya":
                        target_dt = None
                        tgl_faktur_dt_val = None

                        if pd.notna(inv_row.get("Tgl Faktur")):
                            try:
                                tgl_val = str(inv_row["Tgl Faktur"]).lower()
                                for indo, eng in indo_months.items():
                                    if indo in tgl_val:
                                        tgl_val = tgl_val.replace(indo, eng)
                                        break

                                parsed_tf = pd.to_datetime(
                                    tgl_val, errors="coerce", format="mixed"
                                )
                                if pd.notna(parsed_tf):
                                    tgl_faktur_dt_val = parsed_tf.date()
                            except Exception:
                                tgl_faktur_dt_val = None

                        if (
                            flag_giro == "Ya"
                            and "Tanggal JT" in inv_row
                            and pd.notna(inv_row.get("Tanggal JT"))
                        ):
                            target_dt = parse_max_jt_date(
                                inv_row.get("Tanggal JT")
                            )

                        if target_dt is not None and tgl_faktur_dt_val is not None:
                            selisih_hari = abs((target_dt - tgl_faktur_dt_val).days)
                            inv_part.append(f"{selisih_hari}\tHR")
                        elif target_dt is None and tgl_faktur_dt_val is not None:
                            selisih_hari = abs((current_date - tgl_faktur_dt_val).days)
                            inv_part.append(f"{selisih_hari}\tHR")
                        else:
                            inv_part.append("-\tHR")

                    line_str = "\t".join(
                        [str(x) for x in inv_part if x != ""]
                    )

                    if (
                        flag_giro == "Ya"
                        and "Tanggal JT" in inv_row
                        and pd.notna(inv_row["Tanggal JT"])
                    ):
                        tgl_jt_giro = format_excel_date(inv_row["Tanggal JT"])
                        if tgl_jt_giro.strip():
                            line_str += f" ({tgl_jt_giro})"

                    if flag_fraud == "Ya" and "Nama Penjual" in inv_row:
                        if (
                            "FRAUD"
                            in str(inv_row.get("Nama Penjual", "")).upper()
                        ):
                            line_str += " (FRAUD)"

                    note_lines.append(line_str)

            final_note_text = "\n".join(note_lines)

            req_item = {
                "updateCells": {
                    "rows": [
                        {
                            "values": [
                                {
                                    "userEnteredValue": {
                                        "stringValue": formatted_piutang_val
                                        if flag_calc == "Ya"
                                        else ""
                                    },
                                    "note": final_note_text,
                                }
                            ]
                        }
                    ],
                    "fields": "userEnteredValue,note",
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": row_idx - 1,
                        "endRowIndex": row_idx,
                        "startColumnIndex": target_col_idx,
                        "endColumnIndex": target_col_idx + 1,
                    },
                }
            }
            requests.append(req_item)
            total_diisi += 1

        if requests:
            BATCH_SIZE = 300
            for i in range(0, len(requests), BATCH_SIZE):
                chunk = requests[i : i + BATCH_SIZE]
                ss.batch_update({"requests": chunk})
                print(
                    f"--> Mengunggah {len(chunk)} baris ke Google Sheets {prod_suffix.upper()}..."
                )

            print(
                f"--> Total {total_diisi} baris data kosong untuk {prod_suffix.upper()} berhasil diperbarui!"
            )
        else:
            print(
                f"--> Tidak ada data target kosong baru untuk {prod_suffix.upper()}."
            )


if __name__ == "__main__":
    while True:
        try:
            config_load = load_config()
            interval_menit = int(
                config_load.get("AR", "ar_time_interval", fallback=15)
            )
        except Exception:
            interval_menit = 15

        try:
            run_ar_process()
        except Exception as err:
            print(f"--> Terjadi error runtime saat proses berjalan: {err}")

        next_run = datetime.now() + timedelta(minutes=interval_menit)
        jam_berikutnya = next_run.strftime("%H:%M:%S")

        print(
            f"--> Proses selesai dan dalam mode STANDBY. Menunggu {interval_menit} menit (Eksekusi berikutnya pukul {jam_berikutnya})"
        )

        time.sleep(interval_menit * 60)